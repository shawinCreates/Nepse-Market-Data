import os
import pandas as pd
from openpyxl import load_workbook

from .paths import CSV_DIR, COMBINED_EXCEL, CSV_LIST_FILE, ensure_directories


def write_csv_list():
    csv_files = sorted(
        [
            f.name
            for f in CSV_DIR.glob("*.csv")
        ],
        reverse=True,
    )

    with open(
        CSV_LIST_FILE,
        "w",
        encoding="utf-8",
    ) as f:
        for file in csv_files:
            f.write(file + "\n")


def rebuild_workbook():
    ensure_directories()

    csv_files = sorted(
        [
            f.name
            for f in CSV_DIR.glob("*.csv")
        ],
        reverse=True,
    )

    if not csv_files:
        return

    with pd.ExcelWriter(
        COMBINED_EXCEL,
        engine="openpyxl",
    ) as writer:
        for csv_file in csv_files:
            path = CSV_DIR / csv_file

            df = pd.read_csv(path)

            sheet_name = os.path.splitext(
                csv_file
            )[0]

            df.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )

    write_csv_list()


def update_latest_sheet():
    ensure_directories()

    csv_files = sorted(
        [
            f.name
            for f in CSV_DIR.glob("*.csv")
        ],
        reverse=True,
    )

    if not csv_files:
        return

    latest = csv_files[0]

    if not COMBINED_EXCEL.exists():
        rebuild_workbook()
        return

    workbook = load_workbook(
        COMBINED_EXCEL
    )

    sheet_name = latest.replace(
        ".csv",
        ""
    )

    if sheet_name in workbook.sheetnames:
        workbook.remove(
            workbook[sheet_name]
        )

    df = pd.read_csv(
        CSV_DIR / latest
    )

    with pd.ExcelWriter(
        COMBINED_EXCEL,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        writer._book = workbook

        df.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
        )

    write_csv_list()
